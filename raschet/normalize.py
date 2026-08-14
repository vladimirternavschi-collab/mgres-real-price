# -*- coding: utf-8 -*-
"""Слой нормализации.

Каждый источник превращается в опрятную таблицу с колонками:
    дата | ряд | показатель | значение | единица | источник_файл | флаг

Результат: raschet/data/*.csv
Запуск: python normalize.py
"""
import csv, glob, io, re, sys, zipfile, warnings
from datetime import datetime, date
from pathlib import Path

import openpyxl
import pandas as pd

from config import F, DATA, SRC, COMTRADE_GLOB, MOLDELECTRICA_GLOB

warnings.filterwarnings("ignore")

COLS = ["дата", "ряд", "показатель", "значение", "единица", "источник_файл", "флаг"]
LUNI = ["Ianuarie","Februarie","Martie","Aprilie","Mai","Iunie",
        "Iulie","August","Septembrie","Octombrie","Noiembrie","Decembrie"]


def _write(rows, name):
    p = DATA / name
    with open(p, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(COLS)
        w.writerows(rows)
    print(f"  -> {name}: {len(rows)} строк")
    return p


# ---------------------------------------------------------------- НБС месячные
def nbs_monthly(path, ryad, unit):
    """ENE010400 (газ) и ENE010100 (электроэнергия).

    Год в строке 3, месяц в строке 4, значения с строки 5.
    Значение '..' означает «нет данных», НЕ ноль.
    """
    ws = openpyxl.load_workbook(path, data_only=True)[
        openpyxl.load_workbook(path).sheetnames[0]]
    # карта колонка -> (год, месяц)
    colmap = {}
    year = None
    for c in range(2, ws.max_column + 1):
        y = ws.cell(3, c).value
        if y is not None and str(y).strip().isdigit():
            year = int(str(y).strip())
        m = ws.cell(4, c).value
        if year and m and str(m).strip() in LUNI:
            colmap[c] = (year, LUNI.index(str(m).strip()) + 1)
    rows = []
    for r in range(5, ws.max_row + 1):
        ind = ws.cell(r, 1).value
        if not isinstance(ind, str) or not ind.strip():
            continue
        ind = ind.strip()
        # хвост листа - метаданные, не показатели
        if len(ind) > 90 or ind.startswith(("Informatia", "Sursa", "Drept", "Unitatea",
                                            "Perioada", "Cod intern", "Informație",
                                            "Directia", "tel.", "Biroul", "mii m3 standard")):
            continue
        for c, (y, m) in colmap.items():
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and v.strip() in ("", "..", "…", "-")):
                continue  # нет данных - строку не пишем
            try:
                val = float(str(v).replace(",", ".").replace(" ", ""))
            except ValueError:
                continue
            rows.append([f"{y}-{m:02d}-01", ryad, ind, val, unit, path.name, ""])
    return rows


# ---------------------------------------------------------- НБС годовой баланс
def nbs_balance(path):
    ws = openpyxl.load_workbook(path, data_only=True)[
        openpyxl.load_workbook(path).sheetnames[0]]
    colmap, year = {}, None
    for c in range(2, ws.max_column + 1):
        y = ws.cell(3, c).value
        if y is not None and str(y).strip().isdigit():
            year = int(str(y).strip())
        prod = ws.cell(4, c).value
        if year and prod:
            colmap[c] = (year, str(prod).strip())
    rows = []
    for r in range(5, ws.max_row + 1):
        ind = ws.cell(r, 1).value
        if not isinstance(ind, str) or not ind.strip():
            continue
        ind = ind.strip()
        if len(ind) > 90 or ind.startswith(("Informatia", "Sursa", "Drept", "Unitatea",
                                            "Perioada", "Cod intern", "Informație",
                                            "Directia", "tel.", "Biroul")):
            continue
        for c, (y, prod) in colmap.items():
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and v.strip() in ("", "..", "…", "-")):
                continue
            try:
                val = float(str(v).replace(",", ".").replace(" ", ""))
            except ValueError:
                continue
            rows.append([f"{y}-01-01", "НБС ENE020300 энергобаланс",
                         f"{ind} | {prod}", val, "ТДж", path.name, ""])
    return rows


# ------------------------------------------------------------------- Comtrade
def comtrade():
    rows, seen = [], set()
    for p in sorted(SRC.glob(COMTRADE_GLOB)):
        # в выгрузках Comtrade у строк данных на одно поле больше, чем в шапке
        # (висячая запятая). Без index_col=False pandas сдвигает все колонки.
        d = pd.read_csv(p, index_col=False)
        for _, x in d.iterrows():
            freq = x["freqCode"]
            y = int(x["refYear"])
            if freq == "A":
                dt = f"{y}-01-01"
            else:
                dt = f"{y}-{int(x['refMonth']):02d}-01"
            partner = str(x["partnerDesc"])
            key = (dt, freq, partner)
            if key in seen:
                continue
            seen.add(key)
            ryad = ("Comtrade импорт газа 271121 (годовой)" if freq == "A"
                    else "Comtrade импорт газа 271121 (месячный)")
            val = x.get("primaryValue")
            qty = x.get("qty")
            netw = x.get("netWgt")
            altq = x.get("altQty")
            flag = ""
            # 2016: количества битые (123 млн кг против 800-970 в соседних годах)
            if y == 2016 and pd.notna(qty) and float(qty) < 3e8:
                flag = "подозрительно: количество занижено ~в 7 раз"
            if pd.notna(val):
                rows.append([dt, ryad, f"{partner} | стоимость CIF", float(val), "USD", p.name, flag])
            if pd.notna(qty) and float(qty) > 0:
                rows.append([dt, ryad, f"{partner} | количество", float(qty), "кг", p.name, flag])
            if pd.notna(netw) and float(netw) > 0:
                rows.append([dt, ryad, f"{partner} | нетто-вес", float(netw), "кг", p.name, flag])
            if pd.notna(altq) and float(altq) > 0:
                rows.append([dt, ryad, f"{partner} | энергия", float(altq), "ТДж", p.name, flag])
    return rows


# ------------------------------------------------------------- Pink Sheet (WB)
def pink():
    ws = openpyxl.load_workbook(F["pink"], data_only=True)["Monthly Prices"]
    names = {}
    for c in range(2, ws.max_column + 1):
        n = ws.cell(5, c).value
        u = ws.cell(6, c).value
        if n:
            names[c] = (str(n).strip(), str(u or "").strip("() "))
    rows = []
    for r in range(7, ws.max_row + 1):
        per = ws.cell(r, 1).value
        if not per or not re.match(r"^\d{4}M\d{2}$", str(per).strip()):
            continue
        y, m = str(per).strip().split("M")
        for c, (n, u) in names.items():
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and v.strip() in ("…", "..", "")):
                continue
            try:
                val = float(v)
            except (TypeError, ValueError):
                continue
            rows.append([f"{y}-{int(m):02d}-01", "World Bank Pink Sheet", n, val, u,
                         F["pink"].name, ""])
    return rows


# ------------------------------------------------------------------ BAFA (DE)
def bafa():
    ws = openpyxl.load_workbook(F["bafa"], data_only=False)["Imp_Preise"]
    de = {"Januar":1,"Februar":2,"März":3,"April":4,"Mai":5,"Juni":6,"Juli":7,
          "August":8,"September":9,"Oktober":10,"November":11,"Dezember":12}
    rows = []
    cur_years = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.strip().startswith("Jahr"):
            cur_years = {}
            b, g = ws.cell(r, 2).value, ws.cell(r, 7).value
            if isinstance(b, (int, float)):
                cur_years[2] = int(b)   # Menge B, Wert C
            if isinstance(g, (int, float)):
                cur_years[7] = int(g)   # Menge G, Preis H
            continue
        if isinstance(a, str) and a.strip() in de and cur_years:
            m = de[a.strip()]
            if 2 in cur_years:
                y = cur_years[2]
                menge, wert = ws.cell(r, 2).value, ws.cell(r, 3).value
                if isinstance(menge, (int, float)) and isinstance(wert, (int, float)) and menge:
                    rows.append([f"{y}-{m:02d}-01", "BAFA пограничная цена",
                                 "Preis Euro/TJ", wert / menge * 1000, "EUR/ТДж", F["bafa"].name, "вычислено C/B*1000"])
                    rows.append([f"{y}-{m:02d}-01", "BAFA пограничная цена",
                                 "Menge", float(menge), "ТДж", F["bafa"].name, ""])
            if 7 in cur_years:
                y = cur_years[7]
                menge, preis = ws.cell(r, 7).value, ws.cell(r, 8).value
                if isinstance(preis, (int, float)):
                    rows.append([f"{y}-{m:02d}-01", "BAFA пограничная цена",
                                 "Preis Euro/TJ", float(preis), "EUR/ТДж", F["bafa"].name, ""])
                if isinstance(menge, (int, float)):
                    rows.append([f"{y}-{m:02d}-01", "BAFA пограничная цена",
                                 "Menge", float(menge), "ТДж", F["bafa"].name, ""])
    # дубли года из двух блоков - оставляем первое вхождение
    seen, out = set(), []
    for x in rows:
        k = (x[0], x[2])
        if k in seen:
            continue
        seen.add(k)
        out.append(x)
    return out


# ------------------------------------------------------------------------ FRED
def fred():
    rows = []
    для = {"fred_gas": ("Natural Gas Europe", "USD/mmbtu"),
           "fred_brent": ("Crude Oil Brent", "USD/bbl"),
           "fred_hoil": ("Heating Oil NY Harbor", "USD/gal"),
           "fred_dfuel": ("Diesel NY Harbor", "USD/gal")}
    for key, (name, unit) in для.items():
        p = F[key]
        d = pd.read_csv(p)
        vcol = [c for c in d.columns if c != "observation_date"][0]
        for _, x in d.iterrows():
            if pd.isna(x[vcol]):
                continue
            rows.append([str(x["observation_date"])[:10], "FRED", name,
                         float(x[vcol]), unit, p.name, ""])
    return rows


# ----------------------------------------------------------------------- Ember
def ember():
    d = pd.read_csv(F["ember"])
    rows = []
    for _, x in d.iterrows():
        if pd.isna(x["Price (EUR/MWhe)"]):
            continue
        rows.append([str(x["Date"])[:10], "Ember оптовая цена э/э",
                     str(x["Country"]), float(x["Price (EUR/MWhe)"]),
                     "EUR/МВт·ч", F["ember"].name, ""])
    return rows


# ------------------------------------------------------------------ курсы валют
def currencies():
    rows = []
    # ЕЦБ EUR/USD
    d = pd.read_csv(F["ecb"])
    for _, x in d.iterrows():
        if pd.isna(x["OBS_VALUE"]):
            continue
        rows.append([f"{x['TIME_PERIOD']}-01", "ЕЦБ", "USD за 1 EUR",
                     float(x["OBS_VALUE"]), "USD/EUR", F["ecb"].name, ""])
    # НБМ
    for key, name in [("bnm_usd", "MDL за 1 USD"), ("bnm_eur", "MDL за 1 EUR")]:
        p = F[key]
        with open(p, encoding="utf-8-sig", errors="replace") as fh:
            rd = csv.reader(fh, delimiter=";")
            next(rd, None)
            for r in rd:
                if len(r) < 2 or not r[0].strip():
                    continue
                try:
                    dt = datetime.strptime(r[0].strip(), "%d.%m.%Y").date().isoformat()
                    v = float(r[1].strip().strip('"').replace(",", "."))
                except ValueError:
                    continue
                rows.append([dt, "НБМ", name, v, name.split()[0] + "/" + name.split()[-1],
                             p.name, ""])
    # ЦБ ПМР: cp1251, без заголовка: дата,название,код,единиц,курс,цифровой_код
    with open(F["cbpmr"], encoding="cp1251", errors="replace") as fh:
        rd = csv.reader(fh)
        for r in rd:
            if len(r) < 5:
                continue
            try:
                dt = datetime.strptime(r[0].strip(), "%d.%m.%Y").date().isoformat()
                units = float(r[3])
                rate = float(r[4])
            except ValueError:
                continue
            code = r[2].strip()
            rows.append([dt, "ЦБ ПМР", f"рублей ПМР за {units:g} {code}",
                         rate, f"RUP/{code}", F["cbpmr"].name, ""])
    return rows


# ----------------------------------------------------------- Moldelectrica 15'
def moldelectrica():
    frames = []
    for z in sorted(SRC.glob(MOLDELECTRICA_GLOB)):
        with zipfile.ZipFile(z) as zf:
            for n in zf.namelist():
                if not n.lower().endswith(".csv"):
                    continue
                raw = zf.read(n).decode("utf-8", errors="replace")
                d = pd.read_csv(io.StringIO(raw), sep=";")
                d["источник_файл"] = z.name
                frames.append(d)
    big = pd.concat(frames, ignore_index=True)
    big["ts"] = pd.to_datetime(big["Date / Time"], format="%Y-%m-%d %H:%M", errors="coerce")
    big = big.dropna(subset=["ts"])
    total = len(big)
    dedup = big.sort_values(["ts", "источник_файл"]).drop_duplicates(subset=["ts"], keep="first")
    dups = total - len(dedup)
    print(f"  Moldelectrica: строк {total}, уникальных меток {len(dedup)}, дублей {dups}")
    dedup = dedup.sort_values("ts").reset_index(drop=True)
    dedup.to_csv(DATA / "moldelectrica_15min.csv", index=False, encoding="utf-8-sig")
    # анализ пропусков
    full = pd.date_range(dedup["ts"].min(), dedup["ts"].max(), freq="15min")
    missing = sorted(set(full) - set(dedup["ts"]))
    dst = {"2024-03-31", "2025-03-30", "2026-03-29"}
    real = [m for m in missing if m.strftime("%Y-%m-%d") not in dst]
    with open(DATA / "moldelectrica_propuski.txt", "w", encoding="utf-8") as fh:
        fh.write(f"всего меток в непрерывной сетке: {len(full)}\n")
        fh.write(f"уникальных меток в данных: {len(dedup)}\n")
        fh.write(f"дублей удалено: {dups}\n")
        fh.write(f"пропущенных слотов: {len(missing)}\n")
        fh.write(f"из них переходы на летнее время (31.03.2024, 30.03.2025, 29.03.2026): "
                 f"{len(missing) - len(real)}\n")
        fh.write(f"реальных пропусков: {len(real)}\n\n")
        for m in real:
            fh.write(m.strftime("%Y-%m-%d %H:%M") + "\n")
    print(f"  Moldelectrica: пропусков {len(missing)}, из них реальных {len(real)}")
    return len(dedup), dups, len(missing), len(real)


# --------------------------------------------------------------- бюджет ПМР
def pmr_budget():
    rows = []
    # Приложение № 1 - доходы. Итог: строка «ИТОГО», колонка «ВСЕГО» (последняя)
    ws = openpyxl.load_workbook(F["pmr_dohody"], data_only=True).worksheets[0]
    for r in range(1, ws.max_row + 1):
        lbl = ws.cell(r, 2).value
        if isinstance(lbl, str) and lbl.strip().upper() == "ИТОГО":
            vals = [ws.cell(r, c).value for c in range(3, ws.max_column + 1)]
            nums = [v for v in vals if isinstance(v, (int, float))]
            rows.append(["2026-01-01", "Бюджет ПМР 2026", "Доходы республиканского бюджета, ИТОГО ВСЕГО",
                         float(nums[-1]), "рублей ПМР", F["pmr_dohody"].name,
                         f"лист '{ws.title}', строка {r}, колонка ВСЕГО"])
            break
    # Приложение № 2 - расходы. Итог: строка «ИТОГО», колонка «Всего» (колонка D)
    ws = openpyxl.load_workbook(F["pmr_rashody"], data_only=True).worksheets[0]
    for r in range(1, ws.max_row + 1):
        lbl = ws.cell(r, 3).value
        if isinstance(lbl, str) and lbl.strip().upper() == "ИТОГО":
            v = ws.cell(r, 4).value
            rows.append(["2026-01-01", "Бюджет ПМР 2026", "Расходы республиканского бюджета, ИТОГО Всего",
                         float(v), "рублей ПМР", F["pmr_rashody"].name,
                         f"лист '{ws.title}', строка {r}, колонка D 'Всего'"])
            break
    return rows


# ------------------------------------------------------- контракты МГРЭС 2022
# Значения взяты из OCR договоров (Этап 1), сходимость цена x объём = сумма проверена
CONTRACTS = [
    # (id, документ, дата документа, период с, период по, цена $/кВт·ч, объём МВт·ч, сумма $, пункт)
    ("D1", "Договор № 010422/MGRES", "2022-04-29", "2022-05-01", "2022-05-31",
     0.0595, 221352.0, 13170444.00, "п. 4.1 цена, п. 4.2 сумма, Прил. 3 объём"),
    ("D2a", "Доп. соглашение № 1 к 010422/MGRES", "2022-05-31", "2022-05-01", "2022-05-31",
     0.0595, 221352.0, 13170444.00, "п. 1.5 цена; сумма договора после правки 26 481 327,36 за оба месяца"),
    ("D2b", "Доп. соглашение № 1 к 010422/MGRES", "2022-05-31", "2022-06-01", "2022-06-30",
     0.0599, 222218.420, 13310883.36, "п. 1.5 цена; объём Прил. 3 - 222,218420 тыс. МВт·ч"),
    ("D4a", "Договор № 362-22/MGRES", "2022-06-27", "2022-07-01", "2022-08-01",
     0.0599, 245856.0, None, "п. 4.1 цена; Прил. 3 п. 1.1 объём"),
    ("D4b", "Договор № 362-22/MGRES", "2022-06-27", "2022-08-01", "2022-09-01",
     0.0599, 245856.0, None, "п. 4.1 цена; Прил. 3 п. 1.1 объём"),
    ("D5a", "Доп. соглашение № 1 к 362-22/MGRES", "2022-07-01", "2022-07-01", "2022-08-01",
     0.0599, 245855.274, None, "п. 1.2, Прил. 1: 245,855274 тыс. МВт·ч"),
    ("D6a", "Доп. соглашение № 1 к 408-22/MGRES", "2022-09-30", "2022-09-01", "2022-10-01",
     0.0599, None, 13671653.90, "п. 1.3 a) сумма; объём в документе не приведён"),
    ("D6b", "Доп. соглашение № 1 к 408-22/MGRES", "2022-09-30", "2022-10-01", "2022-11-01",
     0.0625, 246290.0, 15393125.00, "п. 1.2 b) цена, п. 1.3 b) сумма, Прил. 1 п. 2.1 объём"),
    ("D7", "Договор № 489-22/MGRES", "2022-12-03", "2022-12-04", "2023-01-01",
     0.0730, 204763.0, 14947691.76, "п. 4.2 a) цена, п. 4.3 a) сумма, Прил. 3 п. 1.1 объём"),
    ("D8", "Доп. соглашение № 1 к 489-22/MGRES", "2022-12-27", "2023-01-01", "2023-02-01",
     0.0730, 251691.741, 18373497.09, "п. 1.2 b) цена, п. 1.3 b) сумма, Прил. 1 объём"),
]


def contracts():
    rows = []
    for cid, doc, ddate, p1, p2, price, vol, total, place in CONTRACTS:
        rows.append([p1, "Контракты Energocom-МГРЭС 2022", f"{cid} | {doc} | цена",
                     price, "USD/кВт·ч", "OCR_Contracts_MGRES_FULL.txt", place])
        if vol is not None:
            rows.append([p1, "Контракты Energocom-МГРЭС 2022", f"{cid} | {doc} | объём",
                         vol, "МВт·ч", "OCR_Contracts_MGRES_FULL.txt", place])
        if total is not None:
            rows.append([p1, "Контракты Energocom-МГРЭС 2022", f"{cid} | {doc} | сумма",
                         total, "USD", "OCR_Contracts_MGRES_FULL.txt", place])
    return rows


# ------------------------------------------------------------- тарифы ANRE HCA
ANRE_HCA = [
    ("2026-02-03", "HCA 44 | цена газа, вход в сети транспортировки, без НДС", 7906, "лей/1000 м³", "HCA nr. 44 din 03.02.2026.pdf, приложение"),
    ("2026-02-03", "HCA 44 | цена газа, выход из сетей транспортировки, без НДС", 8270, "лей/1000 м³", "HCA nr. 44 din 03.02.2026.pdf, приложение"),
    ("2026-02-03", "HCA 44 | цена газа, выход из распределит. сетей высокого давления, без НДС", 8365, "лей/1000 м³", "HCA nr. 44 din 03.02.2026.pdf, приложение"),
    ("2026-02-03", "HCA 44 | цена газа, выход из распределит. сетей среднего давления, без НДС", 9533, "лей/1000 м³", "HCA nr. 44 din 03.02.2026.pdf, приложение"),
    ("2026-02-03", "HCA 44 | цена газа, выход из распределит. сетей низкого давления, без НДС", 13353, "лей/1000 м³", "HCA nr. 44 din 03.02.2026.pdf, приложение"),
    # Дата постановления уточнена 12.08.2026: 25 марта 2026 года, а не 24-е.
    # Имя файла в наборе (24.03.2026) ошибочно. Подтверждено тремя источниками:
    # шапка самого постановления («HOTĂRÂRE nr. 214 din 25 martie 2026»);
    # реестр тарифов на сайте ANRE («Hot. ANRE nr. 214 din 25.03.2026»);
    # пресс-релиз ANRE и сообщение Moldpres от 25.03.2026 о заседании в тот день.
    ("2026-04-01", "HCA 214 от 25.03.2026 | тариф на передачу э/э Moldelectrica, без НДС", 245, "лей/МВт·ч",
     "HCA nr. 214 din 25.03.2026, п. 1 (файл в наборе назван 24.03.2026 - ошибка имени файла)"),
]


def anre_hca():
    return [[d, "Постановления ANRE (HCA)", n, v, u, src.split(",")[0], src]
            for d, n, v, u, src in ANRE_HCA]


# ---------------------------------------------------------- данные из ANRE PDF
ANRE_REPORTS = [
    # (год, показатель, значение, единица, файл, точное место)
    # 2018 - добавлено 12.08.2026 при разборе выброса в сверке двух статслужб.
    # Цифра снята с графика Figura 1 отчёта ANRE за 2018 год.
    (2018, "Закупка э/э у CTE Moldovenească", 2544.0, "млн кВт·ч",
     "Raport anual de activitate   ANRE 2018.pdf",
     "«Figura 1. Evoluția producerii, importului şi procurărilor de energie electrică "
     "în perioada 2001-2018, mil. kWh», линия «CTE Moldovenească, Transnistria», стр. PDF 10"),
    (2018, "Импорт э/э (ANRE)", 956.0, "млн кВт·ч",
     "Raport anual de activitate   ANRE 2018.pdf",
     "«Figura 1», линия «Importul de energie», стр. PDF 10"),
    (2021, "Закупка э/э у CTE Moldovenească", 3445.6, "млн кВт·ч", "Raport 2021 Final 1.pdf", "стр. PDF 21"),
    (2022, "Закупка э/э у CTE Moldovenească", 2706.0, "млн кВт·ч", "Raport privind Activitatea ANRE in anul 2022.pdf", "стр. PDF 17"),
    (2023, "Закупка э/э у CTE Moldovenească", 3278.5, "млн кВт·ч", "Raportul de activitate 2023.pdf", "стр. PDF 17"),
    (2024, "Закупка э/э у CTE Moldovenească", 3068.8, "млн кВт·ч", "Raportul privind Activitatea ANRE în anul 2024.pdf", "стр. PDF 10"),
    (2025, "Закупка э/э у CTE Moldovenească", 0.0, "млн кВт·ч", "Raportul de activitate 2025.pdf", "стр. PDF 10"),

    # ---------------------------------------------------------------------------
    # Добавлено 13.08.2026. Проверялось, есть ли розничная цена до 2020 года.
    # Ответ: есть, но это ДРУГОЙ показатель. Ряд 2020-2025 «Средняя цена поставки
    # э/э потребителям» - фактическая средневзвешенная за год по всем потребителям,
    # одна таблица в отчёте за 2025 год. А ниже - утверждённый тариф конкретного
    # поставщика с конкретной даты. Компаний две, цифры у них разные, год неполный.
    # В один ряд с ними НЕ сводить. На сайте используются только как оговорка.
    (2012, "Тариф поставки э/э, центр (RED Union Fenosa), с мая 2012", 152.0, "бань/кВт·ч",
     "Raport anual de activitate a ANRE pentru anul 2012_0.pdf",
     "стр. PDF 30: «tariful mediu de furnizare a energiei electrice pentru consumatorii "
     "deserviţi de Î.C.S. RED Union Fenosa S.A. din mai 2012 a fost determinat în mărime "
     "de 152 bani/kWh»"),
    (2012, "Тариф поставки э/э, север (RED Nord), с мая 2012", 167.0, "бань/кВт·ч",
     "Raport anual de activitate a ANRE pentru anul 2012_0.pdf",
     "стр. PDF 30: «tariful mediu de livrare pentru consumatorii deserviţi de S.A. RED Nord "
     "- în mărime de 167 bani/kWh»"),
    (2017, "Тариф поставки э/э, центр (Gas Natural Fenosa), с марта 2017", 192.0, "бань/кВт·ч",
     "Raport anual de activitate a ANRE in anul 2017.pdf",
     "стр. PDF 98: «prețul mediu de furnizare a energiei electrice pentru consumatorii "
     "finali ai Î.C.S. Gas Natural Fenosa Furnizare Energie S.R.L. din martie 2017 a fost "
     "stabilit în mărime de 192 bani/kWh»"),
    (2017, "Тариф поставки э/э, север (FEE Nord), с марта 2017", 202.0, "бань/кВт·ч",
     "Raport anual de activitate a ANRE in anul 2017.pdf",
     "стр. PDF 98: «consumatorilor finali ai S.A. Furnizarea Energiei Electrice Nord "
     "... 202 bani/kWh»"),
    (2018, "Тариф поставки э/э, центр (Gas Natural Fenosa), с 01.07.2018", 173.0, "бань/кВт·ч",
     "Raport anual de activitate   ANRE 2018.pdf",
     "стр. PDF 107: «din 1 iulie 2018 a fost stabilit în mărime de 173 bani/kWh, "
     "cu 10.1 % mai mic față de prețul anterior»"),
    (2018, "Тариф поставки э/э, север (FEE Nord), с 01.07.2018", 179.0, "бань/кВт·ч",
     "Raport anual de activitate   ANRE 2018.pdf",
     "стр. PDF 107: «pentru consumatorii S.A. Furnizarea Energiei Electrice Nord a fost "
     "stabilit în mărime de 179 bani/kWh»"),
    (2019, "Тариф поставки э/э, центр (Premier Energy), с 23.08.2019", 177.0, "бань/кВт·ч",
     "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf",
     "стр. PDF 109: «prețul mediu de furnizare a energiei electrice pentru consumatorii "
     "Î.C.S. Premier Energy S.R.L. din 23 august 2019 a fost stabilit în mărime de "
     "177 bani/kWh»"),
    (2005, "Средняя цена закупки газа (правый берег)", 76.1, "USD/1000 м³", "Raport anual de activitate a ANRE in anul 2017.pdf", "Табл. 12, стр. PDF 24"),
    (2010, "Средняя цена закупки газа (правый берег)", 250.1, "USD/1000 м³", "Raport anual de activitate a ANRE in anul 2017.pdf", "Табл. 12, стр. PDF 24"),
    (2015, "Средняя цена закупки газа (правый берег)", 256.0, "USD/1000 м³", "Raport anual de activitate a ANRE in anul 2017.pdf", "Табл. 12, стр. PDF 24"),
    (2016, "Средняя цена закупки газа (правый берег)", 193.5, "USD/1000 м³", "Raport anual de activitate a ANRE in anul 2017.pdf", "Табл. 12, стр. PDF 24"),
    (2017, "Средняя цена закупки газа (правый берег)", 162.05, "USD/1000 м³", "Raport 2021 Final 1.pdf", "Табл. 13, стр. PDF 42"),
    (2018, "Средняя цена закупки газа (правый берег)", 217.5, "USD/1000 м³", "Raport 2021 Final 1.pdf", "Табл. 13, стр. PDF 42"),
    (2019, "Средняя цена закупки газа (правый берег)", 233.7, "USD/1000 м³", "Raport 2021 Final 1.pdf", "Табл. 13, стр. PDF 42"),
    (2020, "Средняя цена закупки газа (правый берег)", 148.87, "USD/1000 м³", "Raport 2021 Final 1.pdf", "Табл. 13, стр. PDF 42"),
    (2021, "Средняя цена закупки газа (правый берег)", 309.5, "USD/1000 м³", "Raport 2021 Final 1.pdf", "Табл. 13, стр. PDF 42"),
    (2022, "Средняя цена закупки газа (правый берег)", 841.6, "USD/1000 м³", "Raport privind Activitatea ANRE in anul 2022.pdf", "стр. PDF 40"),
    (2023, "Средняя цена закупки газа (правый берег)", 783.9, "USD/1000 м³", "Raportul de activitate 2023.pdf", "Табл. 14, стр. PDF 43"),
    (2025, "Поставка газа Moldovagaz → Tiraspoltransgaz, февраль-декабрь", 689.2, "млн м³", "Raportul de activitate 2025.pdf", "стр. PDF 51"),
    (2010, "Контрактная цена э/э CERS Moldovenească (до пересмотра)", 4.69, "цент/кВт·ч", "Raport final 2010_31.03.11.doc", "раздел о тарифах, январь 2010"),
    (2010, "Контрактная цена э/э CERS Moldovenească (после пересмотра, +24%)", 5.83, "цент/кВт·ч", "Raport final 2010_31.03.11.doc", "раздел о тарифах, январь 2010"),
    (2020, "Средняя цена поставки э/э потребителям без НДС", 175.9, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2021, "Средняя цена поставки э/э потребителям без НДС", 154.0, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2022, "Средняя цена поставки э/э потребителям без НДС", 242.8, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2023, "Средняя цена поставки э/э потребителям без НДС", 290.9, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2024, "Средняя цена поставки э/э потребителям без НДС", 228.0, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2025, "Средняя цена поставки э/э потребителям без НДС", 366.6, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2020, "Средняя цена закупки э/э лицензиатами", 97.3, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2021, "Средняя цена закупки э/э лицензиатами", 104.1, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2022, "Средняя цена закупки э/э лицензиатами", 173.0, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2023, "Средняя цена закупки э/э лицензиатами", 178.7, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2024, "Средняя цена закупки э/э лицензиатами", 167.8, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2025, "Средняя цена закупки э/э лицензиатами", 238.7, "бань/кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2025, "Закупка э/э лицензиатами всего", 4710.1, "млн кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2024, "Закупка э/э лицензиатами всего", 4626.1, "млн кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2023, "Закупка э/э лицензиатами всего", 4333.1, "млн кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2022, "Закупка э/э лицензиатами всего", 4512.9, "млн кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2021, "Закупка э/э лицензиатами всего", 4591.7, "млн кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    (2020, "Закупка э/э лицензиатами всего", 4269.8, "млн кВт·ч", "Raportul de activitate 2025.pdf", "Табл. 2, стр. PDF 10"),
    # --- добавлено промптом № 3: ряд средней цены закупки уходит вглубь до 2001 года
    (2001, "Средняя цена закупки э/э лицензиатами", 36.36, "бань/кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2005, "Средняя цена закупки э/э лицензиатами", 35.13, "бань/кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2010, "Средняя цена закупки э/э лицензиатами", 75.75, "бань/кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2017, "Средняя цена закупки э/э лицензиатами", 99.04, "бань/кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2018, "Средняя цена закупки э/э лицензиатами", 99.36, "бань/кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2019, "Средняя цена закупки э/э лицензиатами", 105.60, "бань/кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2001, "Закупка э/э лицензиатами всего", 3194.8, "млн кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2005, "Закупка э/э лицензиатами всего", 3359.5, "млн кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2010, "Закупка э/э лицензиатами всего", 3835.7, "млн кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2017, "Закупка э/э лицензиатами всего", 4066.4, "млн кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2018, "Закупка э/э лицензиатами всего", 4178.8, "млн кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    (2019, "Закупка э/э лицензиатами всего", 4301.9, "млн кВт·ч", "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "Табл. 1, стр. PDF 16"),
    # --- ЕДИНСТВЕННАЯ прямая цена МГРЭС в отчётах ANRE после 2010 года
    (2022, "Цена закупки э/э у CTE Moldovenească, декабрь", 73.5, "USD/МВт·ч",
     "Raport privind Activitatea ANRE in anul 2022.pdf",
     "стр. PDF 17: «creșterea prețului de procurare a energiei electrice de la sursa principală "
     "CTE Moldovenească cu circa 37,4 %, ajungând în luna decembrie la valoarea de 73,5 USD/MWh»"),
    (2022, "Диапазон цен закупки э/э на рынке Румынии, нижняя граница", 90.0, "USD/МВт·ч",
     "Raport privind Activitatea ANRE in anul 2022.pdf", "стр. PDF 17"),
    (2022, "Диапазон цен закупки э/э на рынке Румынии, верхняя граница", 250.0, "USD/МВт·ч",
     "Raport privind Activitatea ANRE in anul 2022.pdf", "стр. PDF 17"),
    # --- средняя цена импорта э/э (МГРЭС + Украина, blended)
    (2018, "Средняя цена импорта э/э (blended, вкл. МГРЭС)", 5.16, "цент/кВт·ч",
     "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf",
     "стр. PDF 109: «majorarea prețului mediu de import al energiei electrice de la 5,16 ¢/kWh la 5,38 ¢/kWh»"),
    (2019, "Средняя цена импорта э/э (blended, вкл. МГРЭС)", 5.38, "цент/кВт·ч",
     "RAPORT DE ACTIVITATE 2019 FINAL 30.11.2020.docx.pdf", "стр. PDF 109"),
    # --- газ 2010: независимая точка для сверки с Comtrade
    (2009, "Цена закупки газа Moldovagaz у Газпрома (до 01.01.2010)", 196.6, "USD/1000 м³",
     "Raport final 2010_31.03.11.doc",
     "«majorarea, cu începere din 1 ianuarie 2010, a preţului de procurare a gazelor naturale "
     "de către S.A. „Moldovagaz” de la concernul rus S.A.D. „Gazprom” cu 17,9%, de la 196,6 USD până la 231,7 USD»"),
    (2010, "Цена закупки газа Moldovagaz у Газпрома (с 01.01.2010)", 231.7, "USD/1000 м³",
     "Raport final 2010_31.03.11.doc", "тот же абзац"),
    (2010, "Средний тариф на газ для конечных потребителей", 3542.0, "лей/1000 м³",
     "Raport final 2010_31.03.11.doc",
     "«creştere cu 13,9 % a tarifului mediu la gazele naturale... 3542 lei/1000 m3»"),
    (2010, "Курс, применённый ANRE в тарифном решении", 12.3, "лей/USD",
     "Raport final 2010_31.03.11.doc", "«devalorizarea valutei naţionale... la 12,3 lei/$»"),
    (2009, "Курс, применённый ANRE в предыдущем тарифе", 10.7, "лей/USD",
     "Raport final 2010_31.03.11.doc", "«de la 10,7 lei/$ (rată prevăzută în tariful precedent)»"),
    # --- средняя цена закупки, годы 2011-2016: из отчётов 2013 (OCR), 2014, 2015, 2016
    (2011, "Средняя цена закупки э/э лицензиатами", 81.60, "бань/кВт·ч",
     "Raport anual de activitate a ANRE pentru anul 2013_2.pdf",
     "Табл. 8, стр. PDF ~24; текстового слоя нет, извлечено OCR (tesseract -l ron). "
     "Методика: только операторы распредсетей, без правомочных потребителей"),
    (2012, "Средняя цена закупки э/э лицензиатами", 94.83, "бань/кВт·ч",
     "Raport anual de activitate a ANRE pentru anul 2013_2.pdf", "Табл. 8, OCR"),
    (2013, "Средняя цена закупки э/э лицензиатами", 98.59, "бань/кВт·ч",
     "RAPORT 2014_de_activitate.pdf", "Табл. 1, стр. PDF 5"),
    (2014, "Средняя цена закупки э/э лицензиатами", 105.72, "бань/кВт·ч",
     "RAPORT 2014_de_activitate.pdf", "Табл. 1, стр. PDF 5"),
    (2015, "Средняя цена закупки э/э лицензиатами", 132.91, "бань/кВт·ч",
     "Raport anual de activitate 2015.pdf", "Табл. 1, стр. PDF 9"),
    (2016, "Средняя цена закупки э/э лицензиатами", 113.31, "бань/кВт·ч",
     "Raport anual de activitate_2016.pdf", "Табл. 1, стр. PDF 6"),
    # --- сколько Молдова заплатила Приднестровью за э/э в 2016 году
    (2016, "Платежи Молдовы за э/э МГРЭС (вторичный источник)", 178.0, "млн $",
     "bne IntelliNews, «Moldova terminates power contract with Transnistria»",
     "ВТОРИЧНЫЙ ИСТОЧНИК, первичного подтверждения нет: «In total, Moldova paid "
     "Transnistria $178mn for the electricity imported in 2016»"),
    # --- тендер 2017 года
    (2017, "Цена победителя тендера 2017 (DTEK Trading, Украина)", 50.2, "USD/МВт·ч",
     "Ministerul Economiei RM, 01.04.2017",
     "контракт Energocom-DTEK на 01.04.2017-31.03.2018; оферты МГРЭС 58,5 и 54,4 отклонены"),
]


def anre_reports():
    return [[f"{y}-01-01", "Годовые отчёты ANRE", n, v, u, f, place]
            for y, n, v, u, f, place in ANRE_REPORTS]


# ==================================================== СТАТЕЖЕГОДНИКИ ПМР
# Извлечено через pdfplumber (текстовый слой есть, OCR не нужен).
# Выпуск 2021 - ряды 2016-2020; выпуск 2020 - ряды 2001, 2005, 2010, 2015, 2018, 2019.
# Пересечение 2018/2019 совпало во всех трёх таблицах - это внутренняя сверка изданий.
EZH21 = "statisticheskiy_ezhegodnik_pmr_2021.pdf"
EZH20 = "statisticheskiy_ezhegodnik_pmr_2020.pdf"
EZH19 = "statisticheskiy_ezhegodnik_pmr_2019.pdf"
EZH17 = "statisticheskiy_ezhegodnik_pmr_2017.pdf"
EZH16 = "statisticheskiy_ezhegodnik_pmr_2016.pdf"

# ПРОВЕРКА НА ПЕРЕСЕЧЕНИЯХ ИЗДАНИЙ (12.08.2026).
# Выпуски перекрываются по годам, каждый показывает пятилетнее окно:
#   выпуск 2016 -> 2011-2015, выпуск 2017 -> 2012-2016,
#   выпуск 2019 -> 2014-2018, выпуск 2020 -> 2001/2005/2010/2015,
#   выпуск 2021 -> 2016-2020.
# Сверены все три таблицы (электробаланс, сетевой газ, внешняя торговля)
# по всем перекрывающимся годам: 2012-2015 (изд. 2016 против изд. 2017),
# 2015 (изд. 2016 против изд. 2020), 2016 (изд. 2017 против изд. 2021),
# 2016-2018 (изд. 2019 против изд. 2021).
# РАСХОЖДЕНИЙ НЕ НАЙДЕНО НИ ОДНОГО - все значения совпали дословно.
# Правило «при расхождении берём более позднее издание» осталось неприменённым.
# Отдельно: «Отпущено за пределы республики» за 2018 год = 2543,9 подтверждено
# двумя независимыми изданиями (2019 и 2021). Выброс +16,53% в сверке с НБС
# Молдовы не объясняется ошибкой издания Госстата ПМР.

# --- таблица 8.13 (вып. 2021, стр. PDF 105) и 8.10 (вып. 2020, стр. PDF 105), млн кВт·ч
PMR_ELBALANS = {
    # год: (произведено, получено извне, потреблено внутри, отпущено за пределы, источник, место)
    2001: (3649.9,  None, 2183.1, 1466.8, EZH20, "табл. 8.10, стр. PDF 105"),
    2005: (2995.9, 659.2, 2107.9, 1547.2, EZH20, "табл. 8.10, стр. PDF 105"),
    2010: (5060.6,   1.9, 1671.7, 3390.8, EZH20, "табл. 8.10, стр. PDF 105"),
    2011: (4786.1,   1.6, 1763.0, 3023.1, EZH16, "табл. 8.10, стр. PDF 101"),
    2012: (4884.8,   1.9, 1832.3, 3054.4, EZH16, "табл. 8.10, стр. PDF 101"),
    2013: (3604.2,   1.6, 1643.8, 1962.0, EZH16, "табл. 8.10, стр. PDF 101"),
    2014: (4416.7,   1.8, 1807.6, 2610.9, EZH16, "табл. 8.10, стр. PDF 101"),
    2015: (5111.3,   1.6, 1770.4, 3342.5, EZH20, "табл. 8.10, стр. PDF 105"),
    2016: (4946.3,   1.4, 1604.7, 3343.0, EZH21, "табл. 8.13, стр. PDF 105"),
    2017: (4067.2,   1.7, 1790.4, 2278.5, EZH21, "табл. 8.13, стр. PDF 105"),
    2018: (4434.7,   1.7, 1892.5, 2543.9, EZH21, "табл. 8.13, стр. PDF 105"),
    2019: (4756.0,   1.7, 1901.0, 2856.7, EZH21, "табл. 8.13, стр. PDF 105"),
    2020: (5196.0,   1.7, 1946.0, 3251.7, EZH21, "табл. 8.13, стр. PDF 105"),
}

# --- таблица 4.4.7 «Использование сетевого газа», млн м³
# Это КОММУНАЛЬНАЯ сеть: МГРЭС получает газ напрямую с магистрали и сюда не входит.
PMR_SETEVOY_GAZ = {
    # год: (подано в сеть, отпущено потребителям, в т.ч. населению, источник, место)
    2001: (525.2, 522.2, 196.4, EZH20, "табл. 4.4.7, стр. PDF 82"),
    2005: (464.1, 456.1, 164.8, EZH20, "табл. 4.4.7, стр. PDF 82"),
    2010: (550.3, 538.7, 174.3, EZH20, "табл. 4.4.7, стр. PDF 82"),
    2011: (615.4, 601.7, 184.5, EZH16, "табл. 4.4.7, стр. PDF 78"),
    2012: (625.1, 611.1, 184.1, EZH16, "табл. 4.4.7, стр. PDF 78"),
    2013: (603.3, 590.3, 180.6, EZH16, "табл. 4.4.7, стр. PDF 78"),
    2014: (593.6, 582.6, 180.1, EZH16, "табл. 4.4.7, стр. PDF 78"),
    2015: (569.7, 561.5, 175.4, EZH20, "табл. 4.4.7, стр. PDF 82"),
    2016: (595.0, 586.4, 185.7, EZH21, "табл. 4.4.7, стр. PDF 80"),
    2017: (605.8, 597.5, 190.7, EZH21, "табл. 4.4.7, стр. PDF 80"),
    2018: (635.4, 627.3, 199.1, EZH21, "табл. 4.4.7, стр. PDF 80"),
    2019: (616.4, 608.6, 183.6, EZH21, "табл. 4.4.7, стр. PDF 80"),
    2020: (590.8, 582.9, 180.8, EZH21, "табл. 4.4.7, стр. PDF 80"),
}

# --- таблицы 15.2 и 15.3, строка «топливно-энергетические товары», млн $
# ВНИМАНИЕ: это таможенная категория, а не контрактная цена электроэнергии.
PMR_TEK_TORGOVLYA = {
    # год: (экспорт ТЭ товаров, импорт ТЭ товаров, экспорт всего, импорт всего, источник, место)
    2001: ( 44.7, 149.7,  390.2,  529.2, EZH20, "табл. 15.2 и 15.3, стр. PDF 177"),
    2005: ( 47.2, 173.2,  579.7,  855.6, EZH20, "табл. 15.2 и 15.3, стр. PDF 177"),
    2010: (183.7, 567.3,  584.9, 1294.6, EZH20, "табл. 15.2 и 15.3, стр. PDF 177"),
    2011: (173.8, 775.1,  691.8, 1736.1, EZH16, "табл. 15.2 и 15.3, стр. PDF 168"),
    2012: (193.8, 905.8,  696.6, 1800.2, EZH16, "табл. 15.2 и 15.3, стр. PDF 168"),
    2013: (143.6, 755.0,  586.9, 1661.2, EZH16, "табл. 15.2 и 15.3, стр. PDF 168"),
    2014: (167.0, 757.1,  715.9, 1634.7, EZH16, "табл. 15.2 и 15.3, стр. PDF 168"),
    2015: (227.1, 575.0,  611.1, 1138.3, EZH20, "табл. 15.2 и 15.3, стр. PDF 177"),
    2016: (187.1, 408.8,  530.4,  857.8, EZH21, "табл. 15.2 и 15.3, стр. PDF 175"),
    2017: (106.9, 333.6,  521.6,  927.8, EZH21, "табл. 15.2 и 15.3, стр. PDF 175"),
    2018: (119.5, 447.8,  685.7, 1154.5, EZH21, "табл. 15.2 и 15.3, стр. PDF 175"),
    2019: (144.9, 489.5,  655.9, 1150.8, EZH21, "табл. 15.2 и 15.3, стр. PDF 175"),
    2020: (159.9, 340.7,  632.5, 1009.7, EZH21, "табл. 15.2 и 15.3, стр. PDF 175"),
}


def pmr_ezhegodnik():
    rows = []
    for y, (pr, izv, vn, otp, src, place) in PMR_ELBALANS.items():
        for nm, v in (("Произведено электроэнергии (ПМР)", pr),
                      ("Получено электроэнергии из-за пределов республики (ПМР)", izv),
                      ("Потреблено электроэнергии внутри республики (ПМР)", vn),
                      ("Отпущено электроэнергии за пределы республики (ПМР)", otp)):
            if v is None:
                continue
            rows.append([f"{y}-01-01", "Электробаланс ПМР", nm, v, "млн кВт·ч", src, place])
    for y, (pod, otp, nas, src, place) in PMR_SETEVOY_GAZ.items():
        for nm, v in (("Подано сетевого газа в сеть (ПМР, коммунальная сеть)", pod),
                      ("Отпущено сетевого газа потребителям (ПМР)", otp),
                      ("Отпущено сетевого газа населению (ПМР)", nas)):
            rows.append([f"{y}-01-01", "Сетевой газ ПМР", nm, v, "млн м³", src, place])
    for y, (ex, im, exv, imv, src, place) in PMR_TEK_TORGOVLYA.items():
        for nm, v in (("Экспорт топливно-энергетических товаров (ПМР)", ex),
                      ("Импорт топливно-энергетических товаров (ПМР)", im),
                      ("Экспорт товаров всего (ПМР)", exv),
                      ("Импорт товаров всего (ПМР)", imv)):
            rows.append([f"{y}-01-01", "Внешняя торговля ПМР", nm, v, "млн $", src, place])
    return rows


# ------------------------------- тарифы молдавской генерации, январь 2010, ANRE
# Курс, применённый регулятором в тарифном решении: 12,3 лея за доллар
# (в предыдущем тарифе - 10,7). Источник: Raport final 2010_31.03.11.doc,
# извлечено через `antiword -m UTF-8.txt` (обычная конвертация теряет символ ¢).
KURS_ANRE_2010 = 12.3
KURS_ANRE_2010_PRED = 10.7
MD_TARIFY_2010 = [
    # (название, бань/кВт·ч, тип, дословная привязка)
    ("НГЭС Костешты (Î.S. NHE Costeşti)", 17.00, "генерация", "«tariful la energia electrică produsă de Î.S. „NHE Costeşti”, fiind stabilit în cuantum de 17,00 bani/kWh»"),
    ("ТЭЦ-2 (S.A. CET-2)", 96.60, "генерация", "«produsă de S.A. „CET–2” - 96,60 bani/kWh»"),
    ("ТЭЦ Норд (S.A. CET Nord)", 103.03, "генерация", "«produsă de S.A. „CET Nord” - 103,03 bani/kWh»"),
    ("ТЭЦ-1 (S.A. CET-1)", 131.84, "генерация", "«produsă de S.A. „CET–1” - 131,84 bani/kWh»"),
    ("Розница, RED Union Fenosa, 110 кВ", 95.00, "розница", "«tariful de 95 bani/kWh»"),
    ("Розница, RED Union Fenosa, прочие категории", 133.00, "розница", "«un tarif de 133 bani/kWh»"),
    ("Розница, RED Nord и RED Nord-Vest", 143.00, "розница", "«vor aplica pentru toate categoriile de consumatori tariful de 143 bani/kWh»"),
]

# -------------------- нормативы удельного расхода топлива по трём станциям ПМР
# Приказ Министерства регионального развития, транспорта и связи ПМР,
# зарегистрирован Минюстом ПМР 15.05.2015, «Об утверждении на 2016 год нормативов
# удельного расхода топлива на отпущенную электрическую и тепловую энергию», п. 1.
PMR_NORMATIVY_2016 = [
    ("ЗАО «Молдавская ГРЭС»", 356.1),
    ("ООО «Тиротекс-Энерго»", 303.8),
    ("ООО «Тираспольтрансгаз-Приднестровье», когенерационная станция", 293.0),
]
PMR_NORMATIV_ISTOCHNIK = ("Приказ Министерства регионального развития, транспорта и связи ПМР, "
                          "рег. Минюст ПМР 15.05.2015, «Об утверждении на 2016 год нормативов "
                          "удельного расхода топлива...», п. 1")


# ================================ ПРЯМЫЕ ЦЕНЫ ЗАКУПКИ У МГРЭС ПО ГОДАМ
# Найдены за пределами отчётов ANRE: в них прямая цена МГРЭС встречается только
# дважды (январь 2010 и декабрь 2022). Остальное - тендеры Moldelectrica,
# коммюнике Premier Energy и сообщения Минэкономики Молдовы.
#
# ВАЖНО. Цены из коммюнике Premier Energy включают маржу Energocom
# («Prețul de achiziție include și marja pentru serviciile prestate de furnizorul
# S.A. Energocom»). Это верхняя граница цены самой станции, а не цена у ворот.
MGRES_CENY = [
    # (дата с, дата по, цена USD/МВт·ч, тип, источник, точное место)
    # ---- 2011-2014: контракты распределительной компании напрямую со станцией
    ("2011-04-01", "2012-04-01", 61.0, "контракт",
     "protv.md, 01.04.2011",
     "«potrivit contractului incheiat intre Union Fenosa si centrala de la Cuciurgan, "
     "pretul de achizitie este de 6,1 centi pentru un kilowatt-ora» - вторичный источник, "
     "телеканал; цена за январь-март 2011 не установлена"),
    ("2012-04-02", "2013-01-02", 69.0, "контракт",
     "old.ipn.md, 03.04.2012",
     "«Contractul prevede livrarea energiei electrice la preţul de 6,9 cenţi pentru un "
     "kilowatt, cu 13,2% mai mult decât până la 31 martie» - Gas Natural Fenosa, новый "
     "контракт с Кучурганской станцией"),
    ("2013-01-03", "2014-03-31", 69.0, "контракт",
     "adevarul.ro, 03.01.2013 + radiochisinau.md, 12.07.2013",
     "«Noul acord este pentru o perioadă de un an, iar preţul rămâne cel vechi, de 6,9 cenţi "
     "pentru un kilowatt»; срок подтверждён: «La 31 martie 2014, expiră contractele actuale "
     "de achiziţie a energiei electrice la un preţ de 6,9 cenţi» (adevarul.ro, 25.03.2014)"),
    ("2014-04-01", "2015-03-31", 68.0, "контракт",
     "Ministerul Economiei RM, 01.04.2014 (old.mei.gov.md, архивная копия rise.md)",
     "«s-a reușit reducerea prețului de import a energiei electrice de la 6,9 cenți pentru "
     "un kWh la 6,8 cenți pentru un kWh... Republica Moldova va cumpăra energie electrică "
     "de la aceiași generatori: DTEK Vostokenergo (Ucraina) și Centrala de la Cuciurgan». "
     "ОГОВОРКА: с октября 2014 в цепочку встроен посредник EnergoKapital (Тирасполь), "
     "цена для Молдовы осталась 6,8, но станция получала меньше"),
    # ---- 2019-2020: закупка через SA «Energocom», источник назван в документах
    ("2019-04-01", "2020-03-31", 52.4, "контракт (вкл. маржу Energocom)",
     "rdp.moldelectrica.md, Castigator E_2019 от 15.03.2019 + RED Union Fenosa, 14.03.2019",
     "«Câştigător: SA „Energocom”. Prețul ofertei câştigătoare: Sursa 1 (Ucraina, 15% din "
     "volum) - 0,0644 USD/kWh; Sursa 2 (CERS Moldovenească, 85% din volum) - 0,0524 USD/kWh»; "
     "то же в коммюнике RED Union Fenosa: 85% от CERS Moldovenească по 52,4 USD/MWh, "
     "средневзвешенная по контракту 54,2"),
    ("2020-04-01", "2020-06-30", 49.9, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, таблица «Informația privind prețul de referință», раздел закупок 2020",
     "«01.04.2020 - 30.06.2020 | 0,0499 | SA „Energocom” | CERS Moldovenească»"),
    ("2020-07-01", "2021-03-31", 48.65, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, таблица цен 2020 + rdp.moldelectrica.md, Castigator E_2020 от 22.06.2020",
     "«01.07.2020 - 31.03.2021 | 0,04865 | SA „Energocom” | CERS Moldovenească»; "
     "тендер Moldelectrica: «Pre ul ofertei câ tigătoare: 0,04865 USD/kWh»"),
    ("2015-04-01", "2016-02-29", 67.95, "контракт",
     "moldpres.md, 01.04.2017",
     "«Moldova paid a purchase price of 48,995 dollars per 1 megawatt, compared to "
     "67.95 dollars until 1 March 2016» - вторичный источник, госагентство Молдовы"),
    ("2016-04-01", "2017-03-31", 48.995, "контракт",
     "moldpres.md, 01.04.2017",
     "«Suppliers ... purchased from 1 April 2016 ... 28 per cent cheaper ... 48,995 dollars "
     "per 1 megawatt» - вторичный источник, госагентство Молдовы"),
    ("2017-04-01", "2017-06-06", 58.5, "оферта отклонена",
     "Ministerul Economiei RM, 01.04.2017 (через moldpres, interfax, diez.md)",
     "первая оферта Кучурганской станции на тендере 2017 года; не принята"),
    ("2017-04-01", "2017-06-06", 54.4, "оферта отклонена",
     "Ministerul Economiei RM, 01.04.2017",
     "вторая, сниженная оферта Кучурганской станции; тендер выиграл DTEK Trading с 50,2"),
    ("2017-06-07", "2018-03-31", 45.0, "контракт",
     "Ministerul Economiei si Infrastructurii RM, 05.06.2017",
     "«SA Energocom va procura, începând cu 7 iunie curent, o parte din energia electrică ... "
     "de la Centrala Electrică de la Cuciurgan, la prețul de 45 USD/MWh»"),
    ("2021-04-01", "2022-03-31", 53.5, "контракт",
     "premierenergy.md, 01.04.2021 + rdp.moldelectrica.md, 30.03.2021",
     "«a semnat noul contract ... cu Centrala electrică de la Cuciurgan (ЗАО Молдавская ГРЭС) "
     "... la prețul de 0,0535 USD/kWh»; та же цена - победившая оферта на тендере Moldelectrica"),
    ("2023-01-01", "2023-01-31", 73.5, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Ianuarie 2023", "91,1 % объёма от МГРЭС"),
    ("2023-02-01", "2023-02-28", 73.68, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Februarie 2023", "85,1 % объёма от МГРЭС"),
    ("2023-11-01", "2023-11-30", 66.71, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Noiembrie 2023", "87,96 % объёма от МГРЭС"),
    ("2024-01-01", "2024-01-31", 66.60, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Ianuarie 2024", "68,47 % объёма от МГРЭС"),
    ("2024-02-01", "2024-02-29", 66.68, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Februarie 2024", "82,33 % объёма от МГРЭС"),
    ("2024-03-01", "2024-03-31", 66.60, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Martie 2024", "90,84 % объёма от МГРЭС"),
    ("2024-04-01", "2024-04-30", 66.60, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Aprilie 2024", "97,68 % объёма от МГРЭС"),
    ("2024-05-01", "2024-05-31", 66.60, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Mai 2024", "96,59 % объёма от МГРЭС"),
    ("2024-07-01", "2024-07-31", 66.76, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Iulie 2024", "83,77 % объёма от МГРЭС"),
    ("2024-08-01", "2024-08-31", 66.74, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat August 2024", "89,39 % объёма от МГРЭС"),
    ("2024-09-01", "2024-09-30", 67.08, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Septembrie 2024", "93,10 % объёма от МГРЭС"),
    ("2024-10-01", "2024-10-31", 67.41, "контракт (вкл. маржу Energocom)",
     "premierenergy.md, Comunicat Octombrie 2024", "92,66 % объёма от МГРЭС"),
]

# Тендер 2017 года: цена победителя, для контекста
TENDER_2017_DTEK = 50.2


def mgres_ceny():
    return [[d1, "Прямые цены закупки у МГРЭС", f"Цена э/э МГРЭС, {d1}..{d2} ({tip})",
             v, "USD/МВт·ч", src, place]
            for d1, d2, v, tip, src, place in MGRES_CENY]


def main():
    print("НОРМАЛИЗАЦИЯ")
    _write(nbs_monthly(F["nbs_gaz"], "НБС ENE010400 газ помесячно", "тыс. м³ при 20 °C"),
           "nbs_gaz_mes.csv")
    _write(nbs_monthly(F["nbs_el"], "НБС ENE010100 электроэнергия помесячно", "МВт·ч"),
           "nbs_el_mes.csv")
    _write(nbs_balance(F["nbs_balans"]), "nbs_balans_god.csv")
    _write(comtrade(), "comtrade_gaz.csv")
    _write(pink(), "worldbank_pink.csv")
    _write(bafa(), "bafa.csv")
    _write(fred(), "fred.csv")
    _write(ember(), "ember.csv")
    _write(currencies(), "kursy.csv")
    _write(pmr_budget(), "pmr_budjet.csv")
    _write(contracts(), "kontrakty_mgres.csv")
    _write(anre_hca(), "anre_hca.csv")
    _write(anre_reports(), "anre_otchety.csv")
    _write(pmr_ezhegodnik(), "pmr_ezhegodnik.csv")
    _write(mgres_ceny(), "mgres_ceny.csv")
    moldelectrica()
    print("Нормализация завершена.")


if __name__ == "__main__":
    main()
